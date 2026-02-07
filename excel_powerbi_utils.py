import os
import logging
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment
import requests

logging.basicConfig(level=logging.INFO)

def automate_excel_advanced(data_list: list, output_path: str = 'report.xlsx'):
    """Cria planilha Excel com fórmulas e gráficos automaticamente."""
    if not data_list:
        logging.warning("Lista vazia. Pulando Excel.")
        return
    
    df = pd.DataFrame(data_list)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb['Data']
    
    ws['D1'] = 'Total Value'
    ws['D2'] = f'=SUM(C2:C{len(data_list)+1})'
    ws['D2'].font = Font(bold=True)
    ws['D2'].alignment = Alignment(horizontal='center')
    
    chart = BarChart()
    chart.title = "Values by Item"
    chart.x_axis.title = "Items"
    chart.y_axis.title = "Values"
    if len(data_list) > 0:
        data = Reference(ws, min_col=3, min_row=1, max_row=len(data_list)+1)
        categories = Reference(ws, min_col=2, min_row=2, max_row=len(data_list)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E5")
    
    wb.save(output_path)
    logging.info(f"Relatório Excel salvo em {output_path}")

def automate_powerbi_advanced(data_list: list, dataset_name: str = 'APEX_Dataset'):
    """Publica dataset no Power BI via API."""
    client_id = os.environ.get('POWERBI_CLIENT_ID')
    client_secret = os.environ.get('POWERBI_CLIENT_SECRET')
    tenant_id = os.environ.get('POWERBI_TENANT_ID')
    
    if not all([client_id, client_secret, tenant_id]):
        logging.warning("Credenciais Power BI não configuradas.")
        return
    
    auth_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://analysis.windows.net/powerbi/api/.default'
    }
    response = requests.post(auth_url, data=data)
    if response.status_code != 200:
        logging.error(f"Falha na autenticação: {response.text}")
        return
    
    access_token = response.json()['access_token']
    dataset_url = "https://api.powerbi.com/v1.0/myorg/datasets"
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    dataset_payload = {
        "name": dataset_name,
        "tables": [{"name": "DataTable", "columns": [], "rows": data_list}]
    }
    response = requests.post(dataset_url, json=dataset_payload, headers=headers)
    if response.status_code == 201:
        logging.info(f"Dataset '{dataset_name}' criado no Power BI")
    else:
        logging.error(f"Erro ao criar dataset: {response.text}")

if __name__ == "__main__":
    sample_data = [{'id': 1, 'name': 'Item1', 'value': 100}, {'id': 2, 'name': 'Item2', 'value': 200}]
    automate_excel_advanced(sample_data)
    automate_powerbi_advanced(sample_data)
